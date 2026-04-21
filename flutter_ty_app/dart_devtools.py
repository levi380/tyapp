import json
import openpyxl
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# 读取 perf_data.json
with open("perf_data.json", "r", encoding="utf-8") as f:
    data = json.load(f)

page_switch = [d for d in data if d["type"] == "page_switch"]
frame = [d for d in data if d["type"] == "frame"]
memory = [d for d in data if d["type"] == "memory"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "性能报告"

# 1. 汇总指标
fcp = min((d["timestamp"] for d in frame), default=0)
lcp = max((d["timestamp"] for d in frame), default=0)
tbt = sum(d["total"] for d in frame)
avg_fps = sum(1000 / d["total"] for d in frame if d["total"] > 0) / len(frame) if frame else 0
max_memory = max((d["used"] for d in memory), default=0)
avg_frame = sum(d["total"] for d in frame) / len(frame) if frame else 0

summary_headers = ["FCP(ms)", "LCP(ms)", "TBT(ms)", "平均FPS", "内存峰值(Byte)", "平均帧耗时(ms)"]
summary_values = [fcp, lcp, tbt, round(avg_fps, 2), max_memory, round(avg_frame, 2)]
ws.append(summary_headers)
ws.append(summary_values)
ws.append([])

# 2. 原始数据表
headers = ["时间戳", "类型", "页面", "Build(ms)", "Raster(ms)", "Total(ms)", "已使用内存(Byte)"]
ws.append(headers)

for d in data:
    ws.append([
        d.get("timestamp", ""),
        d.get("type", ""),
        d.get("page", ""),
        d.get("build", ""),
        d.get("raster", ""),
        d.get("total", ""),
        d.get("used", "")
    ])

# 自动调整列宽
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value is not None:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

# 3. 图表区域
start_row = 4

# 内存趋势折线图
if memory:
    mem_chart = LineChart()
    mem_chart.title = "内存使用趋势"
    mem_chart.y_axis.title = "内存(Byte)"
    mem_chart.x_axis.title = "采样点"
    mem_col = headers.index("已使用内存(Byte)") + 1
    ts_col = headers.index("时间戳") + 1
    data_ref = Reference(ws, min_col=mem_col, min_row=start_row+1, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=ts_col, min_row=start_row+1, max_row=ws.max_row)
    mem_chart.add_data(data_ref, titles_from_data=False)
    mem_chart.set_categories(cats_ref)
    ws.add_chart(mem_chart, "J2")

# 帧耗时分布柱状图
if frame:
    frame_chart = BarChart()
    frame_chart.title = "帧耗时分布"
    frame_chart.y_axis.title = "耗时(ms)"
    frame_chart.x_axis.title = "采样点"
    total_col = headers.index("Total(ms)") + 1
    ts_col = headers.index("时间戳") + 1
    data_ref = Reference(ws, min_col=total_col, min_row=start_row+1, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=ts_col, min_row=start_row+1, max_row=ws.max_row)
    frame_chart.add_data(data_ref, titles_from_data=False)
    frame_chart.set_categories(cats_ref)
    ws.add_chart(frame_chart, "J20")

# 帧耗时构成饼图（Build vs Raster）
if frame:
    total_build = sum(d.get("build", 0) for d in frame)
    total_raster = sum(d.get("raster", 0) for d in frame)
    pie_row = ws.max_row + 2
    ws.append([])
    ws.append(["指标", "耗时总和"])
    ws.append(["Build(ms)", total_build])
    ws.append(["Raster(ms)", total_raster])

    pie_chart = PieChart()
    pie_chart.title = "帧耗时构成占比"
    labels = Reference(ws, min_col=1, min_row=pie_row+1, max_row=pie_row+2)
    data_ref = Reference(ws, min_col=2, min_row=pie_row+1, max_row=pie_row+2)
    pie_chart.add_data(data_ref, titles_from_data=False)
    pie_chart.set_categories(labels)
    ws.add_chart(pie_chart, "J38")

# 4. 新增空 sheet 占位（中文）
wb.create_sheet("渲染耗时")
wb.create_sheet("CPU-内核分析")
wb.create_sheet("网络-接口性能")

# 保存 Excel
wb.save("perf_report.xlsx")
